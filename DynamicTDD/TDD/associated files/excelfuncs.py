from openpyxl import Workbook
from openpyxl import load_workbook
# import openpyxl 

def readWorkbook(workbook):
    try:
        print("workbook:", workbook)
        wb = load_workbook(workbook)
        return wb
    except Exception as e:
        print (e)

def getAvailableSheets(wb):
    try:
        # global wb 
        # wb = load_workbook(workbook)
        sheets = wb.sheetnames
        return sheets
    except Exception as e:
        print (e)
        return 0

def getCurrentSheet(wb):
    try:
        # global wb 
        # wb = load_workbook(workbook)
        sheet = wb.active
        # return sheet.title
        print(sheet.title)
        return sheet
    except Exception as e:
        print (e)
        return 0

def getColumnHeaders(ws):
    try:
        # wb = load_workbook(workbook)
        # global wb
        # headers = [c.value for c in next(wb['sheet_name'].iter_rows(min_row=1, max_row=1))]
        # ws = wb.active
        rows = ws.max_row
        columns = ws.max_column
        # headers = [c[0].value for c in ws.iter_rows(min_row=1, max_row=rows)]
        headers = [c[0].value for c in ws.iter_cols(min_row=1, max_col=columns, max_row=1)]
        return headers
    except Exception as e:
        print (e)
        return 0

def createColNameDict(ws):
    try:
        ColNames = {}
        Current  = 0
        for COL in ws.iter_cols(1, ws.max_column):
            ColNames[COL[0].value] = Current
            Current += 1
        return ColNames
    except Exception as e:
        print (e)
        return 0
        


def getColumnValuesByName(ws, colname):
    try:
        ColNames = {}
        Current  = 0
        for COL in ws.iter_cols(1, ws.max_column):
            ColNames[COL[0].value] = Current
            Current += 1
        ColValues = []
        for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ColValues.append(row_cells[ColNames[colname]].value)
        return ColValues
    except Exception as e:
        print (e)
        return 0


def getColumnValuesByAnotherCol(ws, colname, othercol):
    try:
        ColNames = {}
        Current  = 0
        for COL in ws.iter_cols(1, ws.max_column):
            ColNames[COL[0].value] = Current
            Current += 1
        ColValues = {}
        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
            ColValues[row_cells[ColNames[colname]].value] = row_cells[ColNames[othercol]].value
        return ColValues
    except Exception as e:
        print (e)
        return 0

def getColumnValueByAnotherVal(ws, firstcol, secondcol, otherval):
    
    print ("First col:", firstcol,", Second col:", secondcol, ",other value:", otherval)
    ColNames = createColNameDict(ws)
    ColValues = []
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row_cells[ColNames[secondcol]].value ==  otherval :
            ColValues.append(row_cells[ColNames[firstcol]].value)
    return ColValues


def getColumnValueByMultipleVals(ws, firstcol, **kwargs):
    # print ("First col:", firstcol)
    ColNames = createColNameDict(ws)
    conditions = []
    ColValues = []
    
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        conditions = [True if row_cells[ColNames[key]].value == val else False for key, val in kwargs.items()]
        # print ("Conditions:", conditions)
        if all(conditions):
            ColValues.append(row_cells[ColNames[firstcol]].value)
            print ("ColValues:", ColValues)
        conditions.clear()
    return ColValues


